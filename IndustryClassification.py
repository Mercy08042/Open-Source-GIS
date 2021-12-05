'''
    Access数据读取，将企业类型代码数据以列的形式写入excel
    基于匹配字典（.xlsx）与Excel进行匹配，并对Excel添加 产业类型 列数据
    直接处理csv文件，以dicList为媒介，得到新的包含产业类型的文件
    date: 2021/9/11 -- 2021/11/4
    author: Mercy WYM
'''

import csv
import codecs
from re import L
import openpyxl
import pypyodbc
import pandas as pd

########## MDB ##########
def connectMDB(dbName,mdbPath):
    connStr = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ= '+ mdbPath + '\\' + dbName
    conn = pypyodbc.win_connect_mdb(connStr)
    return conn

def getMdbData(tableName,conn):
    sqlString = u'select 行业类别 from ' + tableName # u 中文查询防止出错
    cursor = conn.cursor()
    cursor.execute(sqlString)
    categories = cursor.fetchall()
    cursor.close()
    if conn:
        conn.close()
    return categories

########## EXCEL ##########
# 向excel写入Access数据库得到的行业代码前两位
def  writeID(fileName,categories,valueColumn): 
    wb = openpyxl.load_workbook(fileName)
    ws = wb.active
    ws = wb['Sheet']
    ws.cell(row = 1,column = valueColumn).value = '产业代码'
    print(len(categories))
    for index in range(len(categories)):
        print(str(categories[index]))
        code = str(categories[index])[1:3] # 序列切片
        ws.cell(row = index + 2, column = valueColumn).value = code
    wb.save(fileName)

# read excel data restored relationship between 行业代码/行业名称 with 产业类型
def readRelationExcel(rTableName,sheet):
    wb = openpyxl.load_workbook(rTableName)
    sheet = wb.get_sheet_by_name(sheet)
    relationTable = {} # 字典存储行业代码和产业类型关系
    #columnTitle = []
    for row in sheet.rows: # 顺便读取列名
        temp = []
        for cell in range(len(row)):
            temp.append(row[cell].value)
        relationTable[temp[0]] = temp[1]
    return relationTable

# 基于 对照表 写入 '产业分类', 由于要写入数据的 excel 不同，所以需要给定 字典键名所在列号，字典键值所在列号
def writeMyClassification(fileName,rTable,keyColumn,valueColumn):
    wb = openpyxl.load_workbook(fileName)
    ws = wb.active
    ws = wb['Sheet']
    ws.cell(row = 1,column = valueColumn).value = '产业分类'
    for index in range(2,ws.max_row + 1):
        key = ws.cell(row = index, column = keyColumn).value # key 要么是 id  要么是 行业名称
        if key != None and int(key) in rTable.keys():
            if keyColumn == 5:
                myClass = rTable[key]
            else:
                myClass = rTable[int(key)] #键名匹配
            ws.cell(row = index, column = valueColumn).value = myClass
    wb.save(fileName)

########## CSV ##########
def csvRead(initialcsvPath):
    with codecs.open(initialcsvPath,'r','utf-8') as f:
        data = csv.DictReader(f,delimiter ='\t')
        for row in data:
            category = row['cate1']
            if category != None and category in rTableS2.keys():
                row['industry_type'] = rTableS2[category]
            #print(row)
            myDict = {}
            for item in fieldName:
                myDict[item] = row[item]
            #print(myDict)
            myDictList.append(myDict)
    return myDictList

def csvWrite(finalcsvPath,myDictList,fieldName):
    with codecs.open(finalcsvPath,'w','utf-8') as f:
        writer = csv.DictWriter(f,fieldnames=fieldName)
        writer.writeheader()
        writer.writerows(myDictList)
        
dataRoot = r'E:\Buffalo\IndustryClassification'
rTablePath = r'E:\Buffalo\IndustryClassification\relationTable.xlsx'
initialcsvPath = r'E:\Buffalo\IndustryClassification\firm_2015.csv'
finalcsvPath = r'E:\Buffalo\IndustryClassification\final_firm_2015.csv'
timeListIIE = ['1998','1999','2000','2001','2002','2003','2004','2005','2006','2007','2008','2009','2010','2011','2012','2013']
timeListITBC = ['2005','2010','2015'] 
pathList = ['InitialIIE','ITBC','SpiderWeb']

rTableS1 = readRelationExcel(rTablePath,'Sheet1') #对照表1
rTableS2 = readRelationExcel(rTablePath,'Sheet2') #对照表2

fieldName = ['name','company_type','address','statu','start_date','approved_time','cate1','city','province','lat_wgs','lng_wgs','industry_type']
myDictList = []
myDictList = csvRead(initialcsvPath)
print(myDictList)
csvWrite(finalcsvPath,myDictList,fieldName)

for time in timeListIIE:
    mdbPath = dataRoot + '\\' + pathList[0]
    mdbName = time + '.mdb'
    conn = connectMDB(mdbName,mdbPath)
    categories = getMdbData(str(time),conn)
    exlPath = dataRoot + '\\' + pathList[2] + '\\' +  time + '.xlsx'
    writeID(exlPath,categories,4)
    writeMyClassification(exlPath,rTableS1,4,5)

for time in timeListITBC:
    fileName = dataRoot + '\\' + pathList[1] + '\\' + time + '.xlsx'
    writeMyClassification(fileName,rTableS2,5,10)

'''
# txtReader
header = f.readline()
    print(header)
    lines = f.readlines()
    splitLines = []
    for line in lines:
        items = line.split('\t')
        splitLines.append(items)
'''
